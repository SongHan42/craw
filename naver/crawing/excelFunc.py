import openpyxl as xl

def set_delivery(delivery):
    wb = xl.load_workbook('./naver/crawing/excel/delivery.xlsx')
    for sheet_nm in wb.sheetnames:
        sheet = wb[sheet_nm]
        for row_data in sheet.iter_rows(min_row=2):
            delivery[row_data[1].value] = row_data[0].value;
    wb.close()

def set_origin(origin_area):
    wb = xl.load_workbook('./naver/crawing/excel/originarea.xlsx')
    for sheet_nm in wb.sheetnames:
        sheet = wb[sheet_nm]
        for row_data in sheet.iter_rows(min_row=2):
            origin_area[row_data[1].value] = row_data[0].value;
    wb.close()

def set_category(category):
    wb = xl.load_workbook('./naver/crawing/excel/category.xlsx')
    for sheet_nm in wb.sheetnames:
        sheet = wb[sheet_nm]
        for row_data in sheet.iter_rows(min_row = 2):
            if not category.get(row_data[1].value):
                category[row_data[1].value] = {}
            if not category[row_data[1].value].get(row_data[2].value):
                category[row_data[1].value][row_data[2].value] = {}
            if not row_data[3].value:
                category[row_data[1].value][row_data[2].value][""] = row_data[0].value
            elif not category[row_data[1].value][row_data[2].value].get(row_data[3].value):
                category[row_data[1].value][row_data[2].value][row_data[3].value] = {}
            if row_data[3].value and not row_data[4].value:
                category[row_data[1].value][row_data[2].value][row_data[3].value][""] = row_data[0].value
            elif row_data[3].value and row_data[4].value:
                category[row_data[1].value][row_data[2].value][row_data[3].value][row_data[4].value] = row_data[0].value

def save_xl(data, sheet, row_count):
    for row_data in sheet.iter_rows(min_row=2, max_row=2, values_only=True):
        col_name = list(row_data)
    for idx, key in enumerate(col_name):
        if data.get(key):
            sheet.cell(row=row_count, column=idx + 1).value = data[key]