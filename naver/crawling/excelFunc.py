import openpyxl as xl
from ..models import *

def set_delivery(delivery):
    wb = xl.load_workbook('./naver/crawling/excel/delivery.xlsx')
    for sheet_nm in wb.sheetnames:
        sheet = wb[sheet_nm]
        for row_data in sheet.iter_rows(min_row=2):
            delivery[row_data[1].value] = row_data[0].value
    wb.close()

def set_origin(origin_area):
    wb = xl.load_workbook('./naver/crawling/excel/originarea.xlsx')
    for sheet_nm in wb.sheetnames:
        sheet = wb[sheet_nm]
        for row_data in sheet.iter_rows(min_row=2):
            origin_area[row_data[1].value] = row_data[0].value
    wb.close()

def set_category(category):
    wb = xl.load_workbook('./naver/crawling/excel/category.xlsx')
    for sheet_nm in wb.sheetnames:
        sheet = wb[sheet_nm]
        for row_data in sheet.iter_rows(min_row = 2):
            if not category.get(row_data[1].value):
                category[row_data[1].value] = {}
            if not category[row_data[1].value].get(row_data[2].value):
                category[row_data[1].value][row_data[2].value] = {}
            if not row_data[3].value:
                category[row_data[1].value][row_data[2].value][""] = row_data[0].value
            elif not category[row_data[1].value][row_data[2].value].get(row_data[3].value):
                category[row_data[1].value][row_data[2].value][row_data[3].value] = {}
            if row_data[3].value and not row_data[4].value:
                category[row_data[1].value][row_data[2].value][row_data[3].value][""] = row_data[0].value
            elif row_data[3].value and row_data[4].value:
                category[row_data[1].value][row_data[2].value][row_data[3].value][row_data[4].value] = row_data[0].value

# def save_xl(data, sheet, row_count):
#     for row_data in sheet.iter_rows(min_row=2, max_row=2, values_only=True):
#         col_name = list(row_data)
#     for idx, key in enumerate(col_name):
#         if data.get(key):
#             sheet.cell(row=row_count, column=idx + 1).value = data[key]

def db_to_xl(host_url):
    wb = xl.load_workbook('./naver/crawling/excel/naver.xlsx')
    sheet = wb["일괄등록"]
    sheet.delete_rows(3,1000)

    all_product = Product.objects.all()


    for idx, p in enumerate(all_product):
        # b = product.book_set.all()
        # if (b):
        count = idx + 3
        sheet.cell(row = count, column = 2).value = p.category_code
        sheet.cell(row = count, column = 3).value = p.name
        sheet.cell(row = count, column = 4).value = p.state
        sheet.cell(row = count, column = 5).value = p.price
        sheet.cell(row = count, column = 6).value = p.vat
        sheet.cell(row = count, column = 7).value = p.stock_num
        sheet.cell(row = count, column = 8).value = p.option_type
        option_name = p.optionname_set.all() #옵션 네임 없을 때 다 잘 돌아가나??
            # sheet.cell(row = count, column = 9).value = (option_name.name, "\n")
        option_name_str = ""
        option_value_str = ""
        option_price_str = ""
        option_stock_num_str = ""
        for o_name in option_name:
            option_name_str += o_name.name + "\n"
            option = o_name.option_set.all()
            for o in option:
                if hasattr(o, 'optionstocknum'):
                    option_stock_num_str += str(o.optionstocknum.num) + ","
                option_value_str += str(o.value) + ","
                option_price_str += str(o.price) + ","
            option_value_str = option_value_str[:-1] + "\n"
            option_price_str = option_price_str[:-1] + "\n"
        sheet.cell(row = count, column = 9).value = option_name_str[:-1]
        sheet.cell(row = count, column = 10).value = option_value_str
        sheet.cell(row = count, column = 11).value = option_price_str
        sheet.cell(row = count, column = 12).value = option_stock_num_str[:-1]
        direct_input_option = p.directinputoption_set.all()
        direct_input_option_str = ""
        for direct_input_o in direct_input_option:
            direct_input_option_str += direct_input_o.text + "\n"
        sheet.cell(row = count, column = 13).value = direct_input_option_str[:-1]
        
        
        additional_product_name = p.additionalproductname_set.all()
        additional_product_name_str = ""
        additional_p_value_str = ""
        additional_p_price_str = ""
        additional_p_num_str = ""
        for additional_p_name in additional_product_name:
            additional_product_name_str += additional_p_name.name + "\n"
            additional_product_detail = additional_p_name.additionalproductdetail_set.all()
            for additional_p_detail in additional_product_detail:
                additional_p_value_str += additional_p_detail.value + ","
                additional_p_price_str += str(additional_p_detail.price) + ","
                additional_p_num_str += str(additional_p_detail.num) + ","
            additional_p_value_str = additional_p_value_str[:-1] + "\n"
            additional_p_price_str = additional_p_price_str[:-1] + "\n"
            additional_p_num_str = additional_p_num_str[:-1] + "\n"
        sheet.cell(row = count, column = 14).value = additional_product_name_str[:-1]
        sheet.cell(row = count, column = 15).value = additional_p_value_str
        sheet.cell(row = count, column = 16).value = additional_p_price_str
        sheet.cell(row = count, column = 17).value = additional_p_num_str
        if str(p.main_img).find("http://") == -1 or str(p.main_img).find("https://"):
            sheet.cell(row = count, column = 18).value = "http://" + host_url + str(p.main_img)
        else:
            sheet.cell(row = count, column = 18).value = str(p.main_img)
        sub_imgs = p.subimg_set.all()
        sub_img_str = ""
        for sub_img in sub_imgs:
            sub_img_str += str(sub_img.img) + "\n"
        sheet.cell(row = count, column = 19).value = sub_img_str[:-1]
        # 
        # sheet.cell(row = count, column = 20).value = 
        sheet.cell(row = count, column = 20).value = "1"

        sheet.cell(row = count, column = 21).value = p.brand
        sheet.cell(row = count, column = 22).value = p.manufacturer
        if p.manufacturing_date:
            sheet.cell(row = count, column = 23).value = p.manufacturing_date #확인필요
        if p.effective_date:
            sheet.cell(row = count, column = 24).value = p.effective_date #확인필요
        sheet.cell(row = count, column = 25).value = p.origin_code
        sheet.cell(row = count, column = 26).value = p.importer
        
        # AA
        sheet.cell(row = count, column = 27).value = p.is_plural_origin
        sheet.cell(row = count, column = 28).value = p.origin_direct_input
        
        # 미성년자
        # sheet.cell(row = count, column = 29).value = p.origin_direct_input
        # 배송비 템플릿코드
        # sheet.cell(row = count, column = 30).value = 
        # 배송 방법
        # sheet.cell(row = count, column = 31).value = 
        # sheet.cell(row = count, column = 32).value = 택배사코드
        # sheet.cell(row = count, column = 33).value = 배송비유형
        # sheet.cell(row = count, column = 34).value = 기본배송비
        # sheet.cell(row = count, column = 35).value = 배송비 결제방식
        # sheet.cell(row = count, column = 36).value = 조건부무료-상품판매가합계
        # sheet.cell(row = count, column = 37).value = 수량별부과-수량
        # sheet.cell(row = count, column = 38).value = 구간별2구간
        # sheet.cell(row = count, column = 39).value = 구간별3구간수량
        # sheet.cell(row = count, column = 40).value = 구간별3구간배송비
        # sheet.cell(row = count, column = 41).value = 구간별 추가배송비
        # sheet.cell(row = count, column = 42).value = 반품배송비
        # sheet.cell(row = count, column = 43).value = 교환배송비
        # sheet.cell(row = count, column = 44).value = 지역별 차등 배송비
        # sheet.cell(row = count, column = 45).value = 별도설치비
        # sheet.cell(row = count, column = 46).value = p.info_template_code
        sheet.cell(row = count, column = 47).value = p.info_name
        sheet.cell(row = count, column = 48).value = p.info_model_name
        sheet.cell(row = count, column = 49).value = p.info_authorization
        sheet.cell(row = count, column = 50).value = p.info_manufacturer
        after_service = p.afterservice_set.all()
        if (after_service and after_service[0]):
            sheet.cell(row = count, column = 51).value = after_service[0].template_code
            sheet.cell(row = count, column = 52).value = after_service[0].phone_number
            sheet.cell(row = count, column = 53).value = after_service[0].announcement
            sheet.cell(row = count, column = 54).value = after_service[0].seller_specifics
        
        sheet.cell(row = count, column = 55).value = p.pc_instant_discount_value
        sheet.cell(row = count, column = 56).value = p.pc_instant_discount_unit
        sheet.cell(row = count, column = 57).value = p.mobile_instant_discount_value
        sheet.cell(row = count, column = 58).value = p.mobile_instant_discount_unit
        sheet.cell(row = count, column = 59).value = p.multiple_purchase_discount_condition_value
        sheet.cell(row = count, column = 60).value = p.multiple_purchase_discount_condition_unit
        sheet.cell(row = count, column = 61).value = p.multiple_purchase_discount_value
        sheet.cell(row = count, column = 62).value = p.multiple_purchase_discount_unit

        # sheet.cell(row = count, column = 63).value = 상품구매시 포인트 지급 값
        # sheet.cell(row = count, column = 64).value = 상품구매시 포인트 지급 단위
        # sheet.cell(row = count, column = 65).value = 상품구매시 포인트 지급 단위
            # 82 = ISBN
        sheet.cell(row = count, column = 70).value = p.interset_free_installment_month
        sheet.cell(row = count, column = 71).value = p.gift
        sheet.cell(row = count, column = 73).value = p.review_exposure_state
        sheet.cell(row = count, column = 74).value = p.review_non_exposure_reson

        # book = p.book_set.all()
        # if (book and book[0]):
        if hasattr(p, 'book'):
            b = p.book
            sheet.cell(row = count, column = 76).value = b.ISBN
            sheet.cell(row = count, column = 77).value = b.ISSN
            sheet.cell(row = count, column = 78).value = b.is_independent_publication
            sheet.cell(row = count, column = 79).value = b.publication_date
            sheet.cell(row = count, column = 80).value = b.publisher
            sheet.cell(row = count, column = 81).value = b.writer
            sheet.cell(row = count, column = 82).value = b.painter
            sheet.cell(row = count, column = 83).value = b.translator
            sheet.cell(row = count, column = 84).value = b.is_cultural_expenses_income_tax_deduction
    wb.save("./static/naver.xlsx")