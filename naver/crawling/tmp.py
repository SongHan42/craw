import openpyxl as xl

wb = xl.load_workbook('./excel/delivery.xlsx')
ws = wb.active

col_A = ws["A"]
col_B = ws["B"]

# <option value="none">===선택===</option>
text = "<option value=\""
arr = []

for a in col_A:
    text = text + str(a.value) + "\">"
    arr.append(text)
    text = "<option value=\""

for i, b in enumerate(col_B):
    arr[i] = arr[i] + str(b.value) + "</option>"

for i, a in enumerate(arr):
    print(a)

wb.close()


